import os

from openpyxl.workbook import Workbook

from data_model import TOTAL_STR, DATES_STR
from utils import prtInfo, prtResult, increment_date, parse_date


class ExcelHandler:

    def __init__(self, ratio_per_mission: dict, date: str):
        self.ratio_per_mission = ratio_per_mission
        self.date = date

    @staticmethod
    def increment_character(character: str, idx: int) -> str:
        return chr(ord(character) + idx)

    def create_excel_file(self):
        date_fmt = self.date.replace('/', '_')
        filename = f"ratio_{date_fmt}.xlsx"
        prtInfo(f"\nCreating {filename} in {os.getcwd()}...")
        workbook = Workbook()
        sheet = workbook.active

        start_date, end_date = parse_date(self.date)
        # we add one for the total column
        if self.date.count('/') == 2:
            day_range_max = 6
        elif self.date.count('/') == 1:
            day_range_max = int(end_date.day) + 1

        # create table
        for idx, mission in enumerate(self.ratio_per_mission.keys()):

            # create first colum with mission
            sheet.cell(row=idx + 2, column=1, value=mission)

            # increment day
            for i in range(0, day_range_max):
                date = increment_date(start_date, i)
                sheet.cell(row=1, column=i + 2, value=date)
                if date in self.ratio_per_mission[mission][DATES_STR].keys():
                    sheet.cell(row=idx + 2, column=i + 2, value=self.ratio_per_mission[mission][DATES_STR][date])
                else:
                    sheet.cell(row=idx + 2, column=i + 2, value=0)

            if i == day_range_max - 1:
                sheet.cell(row=1, column=i + 2, value="Total")
                sheet.cell(row=idx + 2, column=i + 2, value=self.ratio_per_mission[mission][TOTAL_STR])

        workbook.save(filename)

        prtResult("The file ratio.xlsx has been successfully created! :)")
